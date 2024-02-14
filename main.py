from liCalc import *
from liData import *
from openpyxl import Workbook
from openpyxl.styles import Alignment

header = ['年份', '年干支', '历名', '蔀名', '入蔀年', '月份', '朔', '大余', '小余', '望', '日期', '大余', '小余', '名称', '中气', '日期', '大余', '小余', '名称', '节气', '日期', '大余', '小余']


# liList[i] = [闰, 月序, 月大小, 朔, 望, 中气, 节气]
def getTable(li, year):
	liList = calendar(li, year)
	zqx = li.ssy * 2
	jqx = li.ssy * 2 + 1 + liList[0][-1].pop() * 2  # 节气序
	table = []
	for i in range(len(liList)):
		buming = gz[li.bsgz[li.rbs]]
		month = liList[i][0] + yuefen[liList[i][1]]
		sdy, sxy = liList[i][3]
		sgz = gz[(sdy+li.bsgz[li.rbs]) % 60]
		wdy, wxy = liList[i][4]
		wgz = gz[(wdy+li.bsgz[li.rbs]) % 60]
		wrq = nlrq[(wdy - sdy) % 60]
		qdy, qxy = liList[i][5]
		if qdy == None: qgz, qrq, zqm = None, None, None
		else:
			qgz = gz[(qdy+li.bsgz[li.rbs]) % 60]
			qrq = nlrq[(qdy - sdy) % 60]
			zqm = jieqi[zqx]
			zqx = (zqx + 2) % 24
		jdy, jxy = liList[i][6]
		if jdy == None: jgz, jrq, jqm = None, None, None
		else:
			jgz = gz[(jdy+li.bsgz[li.rbs]) % 60]
			jrq = nlrq[(jdy - sdy) % 60]
			jqm = jieqi[jqx]
			jqx = (jqx + 2) % 24
		table.append([year, ganzhiYear(year), li.lm, buming, li.rbn+1, month, sgz, sdy, sxy, wgz, wrq, wdy, wxy, zqm, qgz, qrq, qdy, qxy, jqm, jgz, jrq, jdy, jxy])
	return table


# 每部历法独立生成表格
def tableToExcel(start_year, end_year, li):
	workbook = Workbook()
	sheet = workbook.active
	sheet.append(header)
	for cell in sheet[1]:
		cell.alignment = Alignment(horizontal='center', vertical='center')
	row = 1
	for year in range(end_year - start_year + 1):
		table = getTable(li, start_year + year)
		for i in range(len(table)):
			for j in range(len(table[i])):
				sheet.cell(row=row+i+1, column=j+1, value=table[i][j])
				sheet.cell(row+i+1, j+1).alignment = Alignment(horizontal='center', vertical='center')
		for k in range(5):
			sheet.merge_cells(start_row=row+1, end_row=row+i+1, start_column=k+1, end_column=k+1)
		sheet.append([])
		row += i + 2
	sheet.freeze_panes = 'A2'
	workbook.save(li.lm + '.xlsx')


# 古六历，自定义时间
# for gll in gllb[2:]:
# 	start_year, end_year = gllb[:2]
# 	tableToExcel(start_year, end_year, gll)

# 颁行历法，设定时间
for bxnf in bxnb:
	start_year, end_year, li = bxnf
	tableToExcel(start_year, end_year, li)

